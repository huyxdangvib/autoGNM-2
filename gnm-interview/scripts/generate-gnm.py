#!/usr/bin/env python3
"""GNM Excel Generator (gnm-interview fork) — reads Build Spec JSON and produces .xlsx.

Usage:
    python3 generate-gnm.py <build-spec.json> [output.xlsx]

Forked from: ../../../gnm/scripts/generate-gnm.py
Divergence: integer levels (1..12) + is_final bool replace the A/B/C/Z letter scheme.
            B2 suffix and filename use "L{level}" / "L{level}F" convention.
            Spec.warnings[] count is stamped near B2 if any overrides were logged.

Build Spec format: see ../references/build-spec-schema.md.
Excel structure ref: gnm/references/part-14-excel-generation-playbook.md.
"""

import json
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# CANONICAL STYLES (Part 13/14)
# ============================================================
HDR_FILL = PatternFill('solid', fgColor='0070C0')
SUB_FILL = PatternFill('solid', fgColor='DDEBF7')
WHT_FILL = PatternFill('solid', fgColor='FFFFFF')
CONTENT_FILL = PatternFill('solid', fgColor='FFFFFF')

TITLE_FONT = Font(name='Myriad Pro', bold=True, color='000000', size=14)
HDR_FONT = Font(name='Myriad Pro', bold=True, color='FFFFFF', size=11)
SUB_FONT = Font(name='Myriad Pro', bold=True, color='000000', size=11)
VAL_FONT = Font(name='Myriad Pro', color='000000', size=11)
ENG_FONT = Font(name='Myriad Pro', color='0563C1', size=11)
WARN_FONT = Font(name='Myriad Pro', italic=True, color='C00000', size=10)

WRAP = Alignment(wrap_text=True, vertical='top', horizontal='left', indent=1)
NO_WRAP = Alignment(wrap_text=False, vertical='top', horizontal='left', indent=1)
NO_WRAP_NI = Alignment(wrap_text=False, vertical='top', horizontal='left')

THIN = Side(style='thin')
MEDIUM = Side(style='medium')
NEG_FMT = '0;(0);0'

WIDTH_SEP = 3
WIDTH_PHAN_DAU = 14
WIDTH_ZONE1 = 14
WIDTH_ZONE2 = 28
WIDTH_CONSO = 28
WIDTH_MO_RONG = 14


# ============================================================
# LEVEL SUFFIX (gnm-interview divergence)
# ============================================================
def level_suffix(level, is_final, lang='en'):
    """Return the parenthetical suffix for B2 / sheet name.

    L1                    -> '(L1)'
    L1, is_final=True, en -> '(L1 — Final)'
    L1, is_final=True, vi -> '(L1 — Cuối)'
    """
    base = f'L{level}'
    if not is_final:
        return f'({base})'
    final_word = 'Cuối' if lang == 'vi' else 'Final'
    return f'({base} — {final_word})'


def filename_suffix(level, is_final):
    """Filename uses compact 'L{n}' or 'L{n}F' (no spaces, no em-dash)."""
    return f'L{level}{"F" if is_final else ""}'


# ============================================================
# VALIDATION
# ============================================================
def validate_spec(spec):
    """Pre-generation validation per Part 14 §8 + integer-level fork rules."""
    errors = []
    layout = spec['layout']
    f_val = layout['f']
    l2 = layout['L2']
    n = layout['n']
    a = layout['a']
    c = layout['c']

    if a < 2:
        errors.append(f"layout.a = {a}, must be >= 2")
    if c < 2:
        errors.append(f"layout.c = {c}, must be >= 2")
    if not (1 <= f_val <= 5):
        errors.append(f"layout.f = {f_val}, must be 1-5")
    if n < 1:
        errors.append(f"layout.n = {n}, must be >= 1")

    items = spec['zone1']['items']
    if len(items) != n:
        errors.append(f"zone1.items length {len(items)} != layout.n {n}")

    z3 = spec['zone3']
    if len(z3) != n:
        errors.append(f"zone3 has {len(z3)} rows, expected {n}")
    for i, row in enumerate(z3):
        if len(row) != f_val:
            errors.append(f"zone3[{i}] has {len(row)} cols, expected {f_val}")

    z4 = spec['zone4']
    if len(z4) != n:
        errors.append(f"zone4 length {len(z4)} != layout.n {n}")

    features = spec['zone2']['features']
    if len(features) != f_val:
        errors.append(f"zone2.features length {len(features)} != layout.f {f_val}")

    z5 = spec.get('zone5', [])
    if z5 and len(z5) > a:
        errors.append(f"zone5 has {len(z5)} rows, exceeds layout.a {a}")

    z9 = spec.get('zone9', [])
    if z9 and len(z9) > c:
        errors.append(f"zone9 has {len(z9)} rows, exceeds layout.c {c}")

    if l2 == 1:
        for i, item in enumerate(items):
            if 'l2' not in item or item.get('l2') is None:
                errors.append(f"zone1.items[{i}] missing 'l2' field (L2=1)")

    code = spec['gnm']['code']
    if not re.match(r'^[A-Z]{3}$', code):
        errors.append(f"gnm.code '{code}' must be exactly 3 uppercase letters")

    # Integer level (1..12) replaces letter scheme
    level = spec['gnm'].get('level')
    if not (isinstance(level, int) and 1 <= level <= 12):
        errors.append(f"gnm.level '{level}' must be an integer 1..12")

    is_final = spec['gnm'].get('is_final')
    if not isinstance(is_final, bool):
        errors.append(f"gnm.is_final '{is_final}' must be a boolean")

    return errors


# ============================================================
# LAYOUT COMPUTATION
# ============================================================
def compute_layout(spec):
    layout = spec['layout']
    f_val = layout['f']
    l2 = layout['L2']
    n = layout['n']
    a = layout['a']
    c = layout['c']

    zone2_start_col = ord('G') + l2
    zone2_cols = [chr(zone2_start_col + k) for k in range(f_val)]
    conso_col = chr(zone2_start_col + f_val)
    sep_col = chr(zone2_start_col + f_val + 1)
    mr_col1 = chr(zone2_start_col + f_val + 2)
    mr_col2 = chr(zone2_start_col + f_val + 3)

    last_item_row = 7 + n
    all_start = 8 + n
    all_end = 7 + n + a
    common_start = 8 + n + a
    common_end = 7 + n + a + c
    end_row = common_end

    return {
        'f': f_val, 'L2': l2, 'n': n, 'a': a, 'c': c,
        'zone2_cols': zone2_cols,
        'conso_col': conso_col,
        'sep_col': sep_col,
        'mr_col1': mr_col1,
        'mr_col2': mr_col2,
        'last_item_row': last_item_row,
        'all_start': all_start,
        'all_end': all_end,
        'common_start': common_start,
        'common_end': common_end,
        'end_row': end_row,
        'b2_merge': f'B2:{conso_col}2',
    }


def col_num(letter):
    return ord(letter) - ord('A') + 1


def write_engine(ws, cell_ref, entry):
    """Write engine entry — plain text or HYPERLINK."""
    if isinstance(entry, dict) and 'sheet' in entry:
        sheet_name = entry['sheet'].replace("'", "''")
        text = entry.get('text', entry['sheet'])
        ws[cell_ref] = f"=HYPERLINK(\"#'{sheet_name}'!B2\", \"{text}\")"
    else:
        ws[cell_ref] = entry


# ============================================================
# PHASE 1: DATA
# ============================================================
def write_data(ws, spec, lo):
    gnm = spec['gnm']
    z1 = spec['zone1']
    z2 = spec['zone2']
    z3 = spec['zone3']
    z4 = spec['zone4']
    lang = spec.get('session', {}).get('lang', 'en')

    # B2: GNM Name + level suffix (merged)
    ws.merge_cells(lo['b2_merge'])
    suffix = level_suffix(gnm['level'], gnm['is_final'], lang)
    ws['B2'] = f"{gnm['name']} {suffix}"

    # Row 4: Headers — negative numbers
    ws['B4'] = -1
    ws['C4'] = -2

    phan_than_cols = ['E', 'F']
    if lo['L2'] == 1:
        phan_than_cols.append('G')
    phan_than_cols.extend(lo['zone2_cols'])
    phan_than_cols.append(lo['conso_col'])

    for i, col_letter in enumerate(phan_than_cols):
        ws[f'{col_letter}4'] = -(i + 1)

    ws[f'{lo["mr_col1"]}4'] = -1
    ws[f'{lo["mr_col2"]}4'] = -2

    # Row 5: Sub-headers
    ws['B5'] = gnm['code']
    ws['C5'] = -1
    ws[f'{lo["conso_col"]}5'] = 'Conso.'
    ws[f'{lo["mr_col1"]}5'] = 'Common'

    # Row 6: Zone headers
    ws['C6'] = -2
    ws['E6'] = 'Object'
    ws[f'{lo["zone2_cols"][0]}6'] = z2['feature_group']
    ws[f'{lo["conso_col"]}6'] = '-'
    ws[f'{lo["mr_col1"]}6'] = '-'

    # Row 7: Zone headers
    ws['C7'] = -3
    ws['E7'] = 'Item'
    ws['F7'] = '-'
    if lo['L2'] == 1:
        ws['G7'] = '-'

    if lo['f'] == 1:
        ws[f'{lo["zone2_cols"][0]}7'] = '-'
    else:
        for k, feat in enumerate(z2['features']):
            ws[f'{lo["zone2_cols"][k]}7'] = feat

    ws[f'{lo["conso_col"]}7'] = '-'
    ws[f'{lo["mr_col1"]}7'] = '-'
    ws[f'{lo["mr_col2"]}7'] = '-'

    # Rows 8..7+n: Zone 1 items + Zone 3 + Zone 4
    for i, item in enumerate(z1['items']):
        row = 8 + i
        if item.get('l1'):
            ws[f'F{row}'] = item['l1']
        if lo['L2'] == 1 and item.get('l2'):
            ws[f'G{row}'] = item['l2']

        for k in range(lo['f']):
            write_engine(ws, f'{lo["zone2_cols"][k]}{row}', z3[i][k])

        write_engine(ws, f'{lo["conso_col"]}{row}', z4[i])

    ws[f'C{lo["all_start"]}'] = 'All'

    z5 = spec.get('zone5', [])
    for ri, z5_row in enumerate(z5):
        row = lo['all_start'] + ri
        if isinstance(z5_row, list):
            for k, val in enumerate(z5_row):
                if val and k < len(lo['zone2_cols']):
                    write_engine(ws, f'{lo["zone2_cols"][k]}{row}', val)
        elif z5_row:
            write_engine(ws, f'{lo["zone2_cols"][0]}{row}', z5_row)

    z6 = spec.get('zone6', [])
    for ri, val in enumerate(z6):
        if val:
            write_engine(ws, f'{lo["conso_col"]}{lo["all_start"] + ri}', val)

    z7 = spec.get('zone7', [])
    for ri, pair in enumerate(z7):
        row = lo['all_start'] + ri
        if isinstance(pair, list):
            if len(pair) > 0 and pair[0]:
                write_engine(ws, f'{lo["mr_col1"]}{row}', pair[0])
            if len(pair) > 1 and pair[1]:
                write_engine(ws, f'{lo["mr_col2"]}{row}', pair[1])
        elif pair:
            write_engine(ws, f'{lo["mr_col1"]}{row}', pair)

    ws[f'B{lo["common_start"]}'] = 'Common'
    ws[f'C{lo["common_start"]}'] = '-'

    z8 = spec.get('zone8', [])
    for ri, val in enumerate(z8):
        if val:
            write_engine(ws, f'{lo["conso_col"]}{lo["common_start"] + ri}', val)

    z9 = spec.get('zone9', [])
    for ri, pair in enumerate(z9):
        row = lo['common_start'] + ri
        if isinstance(pair, list):
            if len(pair) > 0 and pair[0]:
                write_engine(ws, f'{lo["mr_col1"]}{row}', pair[0])
            if len(pair) > 1 and pair[1]:
                write_engine(ws, f'{lo["mr_col2"]}{row}', pair[1])
        elif pair:
            write_engine(ws, f'{lo["mr_col1"]}{row}', pair)

    # Stamp warnings count if any overrides logged
    warnings = spec.get('warnings', [])
    if warnings:
        ws['B1'] = f'⚠ {len(warnings)} rule override{"s" if len(warnings) != 1 else ""} logged'


# ============================================================
# PHASE 2: FORMULAS
# ============================================================
def write_formulas(ws, spec, lo):
    ws['E5'] = '=B5'
    ws['E8'] = '=B5'
    for i in range(lo['n']):
        ws[f'C{8 + i}'] = '=ROW()-ROW($C$7)'

    parent = spec.get('parent')
    if parent and isinstance(parent, dict) and parent.get('backlink'):
        parent_sheet = parent.get('sheet', '')
        ws['A1'] = f'=HYPERLINK("#\'{parent_sheet}\'!A1", "<<")'


# ============================================================
# PHASE 3: MEDIUM BORDERS
# ============================================================
def set_medium_border_outline(ws, min_row, max_row, min_col, max_col):
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            existing = cell.border
            top = MEDIUM if row == min_row else (existing.top if existing else None)
            bottom = MEDIUM if row == max_row else (existing.bottom if existing else None)
            left = MEDIUM if col == min_col else (existing.left if existing else None)
            right = MEDIUM if col == max_col else (existing.right if existing else None)
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def apply_medium_borders(ws, lo):
    end = lo['end_row']
    conso_num = col_num(lo['conso_col'])
    mr1_num = col_num(lo['mr_col1'])
    mr2_num = col_num(lo['mr_col2'])
    set_medium_border_outline(ws, 4, end, 2, 3)
    set_medium_border_outline(ws, 4, end, 5, conso_num)
    set_medium_border_outline(ws, 4, end, mr1_num, mr2_num)


# ============================================================
# PHASE 4: THIN BORDERS
# ============================================================
def set_thin_top_border(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        existing = cell.border
        cell.border = Border(
            top=THIN,
            bottom=existing.bottom,
            left=existing.left,
            right=existing.right
        )


def apply_thin_borders(ws, lo):
    conso_num = col_num(lo['conso_col'])
    mr1_num = col_num(lo['mr_col1'])
    mr2_num = col_num(lo['mr_col2'])

    cell_c = ws.cell(row=lo['all_start'], column=3)
    existing = cell_c.border
    cell_c.border = Border(top=THIN, bottom=existing.bottom,
                           left=existing.left, right=existing.right)

    set_thin_top_border(ws, lo['all_start'], 5, conso_num)
    set_thin_top_border(ws, lo['all_start'], mr1_num, mr2_num)
    set_thin_top_border(ws, lo['common_start'], 2, 3)
    set_thin_top_border(ws, lo['common_start'], 5, conso_num)
    set_thin_top_border(ws, lo['common_start'], mr1_num, mr2_num)


# ============================================================
# PHASE 5: FORMATTING
# ============================================================
def apply_formatting(ws, spec, lo):
    conso_num = col_num(lo['conso_col'])
    mr1_num = col_num(lo['mr_col1'])
    mr2_num = col_num(lo['mr_col2'])

    ws.column_dimensions['A'].width = WIDTH_SEP
    ws.column_dimensions['B'].width = WIDTH_PHAN_DAU
    ws.column_dimensions['C'].width = WIDTH_PHAN_DAU
    ws.column_dimensions['D'].width = WIDTH_SEP
    ws.column_dimensions['E'].width = WIDTH_ZONE1
    ws.column_dimensions['F'].width = WIDTH_ZONE1
    if lo['L2'] == 1:
        ws.column_dimensions['G'].width = WIDTH_ZONE1

    for col_letter in lo['zone2_cols']:
        ws.column_dimensions[col_letter].width = WIDTH_ZONE2
    ws.column_dimensions[lo['conso_col']].width = WIDTH_CONSO
    ws.column_dimensions[lo['sep_col']].width = WIDTH_SEP
    ws.column_dimensions[lo['mr_col1']].width = WIDTH_MO_RONG
    ws.column_dimensions[lo['mr_col2']].width = WIDTH_MO_RONG

    for row in range(1, lo['end_row'] + 1):
        ws.row_dimensions[row].height = 18

    ws['B2'].font = TITLE_FONT
    ws['B2'].fill = WHT_FILL
    ws['B2'].alignment = NO_WRAP_NI

    if spec.get('warnings'):
        ws['B1'].font = WARN_FONT
        ws['B1'].fill = WHT_FILL
        ws['B1'].alignment = NO_WRAP_NI

    if spec.get('parent'):
        ws['A1'].font = ENG_FONT
        ws['A1'].fill = WHT_FILL
        ws['A1'].alignment = NO_WRAP_NI

    ws.sheet_view.showGridLines = False

    header_cols = ['B', 'C', 'E', 'F']
    if lo['L2'] == 1:
        header_cols.append('G')
    header_cols.extend(lo['zone2_cols'])
    header_cols.append(lo['conso_col'])
    header_cols.append(lo['mr_col1'])
    header_cols.append(lo['mr_col2'])

    for col_letter in header_cols:
        cell = ws[f'{col_letter}4']
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = NO_WRAP
        cell.number_format = NEG_FMT

    ws['B5'].font = SUB_FONT
    ws['B5'].fill = WHT_FILL
    ws['B5'].alignment = NO_WRAP

    for ref in ['C5', 'C6', 'C7']:
        ws[ref].font = SUB_FONT
        ws[ref].fill = SUB_FILL
        ws[ref].alignment = NO_WRAP
        ws[ref].number_format = NEG_FMT

    ws['E5'].font = SUB_FONT
    ws['E5'].fill = SUB_FILL
    ws['E5'].alignment = NO_WRAP

    sub5_cols = ['F']
    if lo['L2'] == 1:
        sub5_cols.append('G')
    sub5_cols.extend(lo['zone2_cols'])
    for col_letter in sub5_cols:
        ws[f'{col_letter}5'].fill = SUB_FILL
        ws[f'{col_letter}5'].alignment = NO_WRAP

    ws[f'{lo["conso_col"]}5'].font = SUB_FONT
    ws[f'{lo["conso_col"]}5'].fill = SUB_FILL
    ws[f'{lo["conso_col"]}5'].alignment = NO_WRAP

    ws[f'{lo["mr_col1"]}5'].font = SUB_FONT
    ws[f'{lo["mr_col1"]}5'].fill = SUB_FILL
    ws[f'{lo["mr_col1"]}5'].alignment = NO_WRAP

    ws[f'{lo["mr_col2"]}5'].fill = SUB_FILL
    ws[f'{lo["mr_col2"]}5'].alignment = NO_WRAP

    zone_hdr_cols = ['E', 'F']
    if lo['L2'] == 1:
        zone_hdr_cols.append('G')
    zone_hdr_cols.extend(lo['zone2_cols'])
    zone_hdr_cols.append(lo['conso_col'])
    zone_hdr_cols.append(lo['mr_col1'])
    zone_hdr_cols.append(lo['mr_col2'])

    for col_letter in zone_hdr_cols:
        for r in [6, 7]:
            cell = ws[f'{col_letter}{r}']
            cell.font = SUB_FONT
            cell.fill = SUB_FILL
            cell.alignment = WRAP

    zone1_cols_nums = [5, 6]
    if lo['L2'] == 1:
        zone1_cols_nums.append(7)

    z2_col_nums = [col_num(c) for c in lo['zone2_cols']]
    mr_col_nums = [mr1_num, mr2_num]

    for i in range(lo['n']):
        row = 8 + i
        for col in [2, 3]:
            cell = ws.cell(row=row, column=col)
            cell.fill = CONTENT_FILL
            cell.font = VAL_FONT
            cell.alignment = NO_WRAP

        for col in zone1_cols_nums:
            cell = ws.cell(row=row, column=col)
            cell.font = VAL_FONT
            cell.fill = CONTENT_FILL
            cell.alignment = WRAP

        for col in z2_col_nums:
            cell = ws.cell(row=row, column=col)
            cell.font = VAL_FONT
            cell.fill = CONTENT_FILL
            cell.alignment = WRAP

        cell = ws.cell(row=row, column=conso_num)
        cell.font = ENG_FONT
        cell.fill = CONTENT_FILL
        cell.alignment = WRAP

        for col in mr_col_nums:
            cell = ws.cell(row=row, column=col)
            cell.fill = CONTENT_FILL
            cell.alignment = WRAP

    for ri in range(lo['a']):
        row = lo['all_start'] + ri
        for col in [2, 3]:
            cell = ws.cell(row=row, column=col)
            cell.fill = CONTENT_FILL
            cell.font = VAL_FONT
            cell.alignment = NO_WRAP

        for col in zone1_cols_nums:
            cell = ws.cell(row=row, column=col)
            cell.fill = CONTENT_FILL
            cell.alignment = WRAP

        for col in z2_col_nums:
            cell = ws.cell(row=row, column=col)
            cell.font = ENG_FONT
            cell.fill = CONTENT_FILL
            cell.alignment = WRAP

        cell = ws.cell(row=row, column=conso_num)
        cell.font = ENG_FONT
        cell.fill = CONTENT_FILL
        cell.alignment = WRAP

        for col in mr_col_nums:
            cell = ws.cell(row=row, column=col)
            cell.font = ENG_FONT
            cell.fill = CONTENT_FILL
            cell.alignment = WRAP

    for ri in range(lo['c']):
        row = lo['common_start'] + ri
        for col in [2, 3]:
            cell = ws.cell(row=row, column=col)
            cell.fill = CONTENT_FILL
            cell.font = VAL_FONT
            cell.alignment = NO_WRAP

        for col in zone1_cols_nums + z2_col_nums:
            cell = ws.cell(row=row, column=col)
            cell.fill = CONTENT_FILL
            cell.alignment = WRAP

        cell = ws.cell(row=row, column=conso_num)
        cell.font = ENG_FONT
        cell.fill = CONTENT_FILL
        cell.alignment = WRAP

        for col in mr_col_nums:
            cell = ws.cell(row=row, column=col)
            cell.font = ENG_FONT
            cell.fill = CONTENT_FILL
            cell.alignment = WRAP


# ============================================================
# MAIN
# ============================================================
def generate_gnm(spec_path, output_path=None):
    with open(spec_path, 'r', encoding='utf-8') as f:
        spec = json.load(f)

    errors = validate_spec(spec)
    if errors:
        print("❌ Build Spec validation failed:")
        for e in errors:
            print(f"   • {e}")
        sys.exit(1)

    print("✅ Build Spec validation passed")

    for zn in ['zone5', 'zone6', 'zone7', 'zone8', 'zone9']:
        if not spec.get(zn):
            print(f"⚠️  {zn} is empty — All/Common cluster will have blank {zn} rows")

    lo = compute_layout(spec)
    gnm = spec['gnm']
    fsuf = filename_suffix(gnm['level'], gnm['is_final'])

    if not output_path:
        output_path = f'{gnm["code"]}-{fsuf}-gnm.xlsx'

    print(f"📐 Layout: f={lo['f']}, L2={lo['L2']}, n={lo['n']}, "
          f"a={lo['a']}, c={lo['c']}")
    print(f"   Level: {gnm['level']} (is_final={gnm['is_final']}) → suffix '{fsuf}'")
    print(f"   Zone 2 cols: {lo['zone2_cols']}, Conso: {lo['conso_col']}, "
          f"Mở rộng: {lo['mr_col1']}-{lo['mr_col2']}")

    wb = Workbook()
    ws = wb.active
    ws.title = gnm.get('sheet_name', f'{gnm["code"]} ({fsuf})')[:31]

    print("\nPhase 1: Writing data...")
    write_data(ws, spec, lo)

    print("Phase 2: Writing formulas...")
    write_formulas(ws, spec, lo)

    print("Phase 3: Applying medium borders...")
    apply_medium_borders(ws, lo)

    print("Phase 4: Applying thin borders...")
    apply_thin_borders(ws, lo)

    print("Phase 5: Applying formatting...")
    apply_formatting(ws, spec, lo)

    wb.save(output_path)
    print(f"\n✅ GNM Excel saved: {output_path}")
    print(f"   Sheet: {ws.title}")

    return output_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 generate-gnm.py <build-spec.json> [output.xlsx]")
        sys.exit(1)

    spec_file = sys.argv[1]
    out_file = sys.argv[2] if len(sys.argv) > 2 else None
    generate_gnm(spec_file, out_file)
